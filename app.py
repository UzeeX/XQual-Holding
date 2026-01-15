import os, textwrap, pandas as pd, re, io

app_code = r'''
import re
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt

st.set_page_config(page_title="Cross-Qualifying Ticker Finder", layout="wide")

# -----------------------------
# Constants / Matching
# -----------------------------
META_COLS_DEFAULT = ["Symbol", "Exchange", "Compagny", "Sector", "Rank"]
KNOWN_SUFFIXES = (".TO", ".V", ".CN", ".NE", ".TSX", ".TSXV")

def normalize_ticker(raw: str) -> str:
    """Normalize user-entered tickers/symbols to a comparable key."""
    if raw is None:
        return ""
    s = str(raw).strip().upper()
    if not s:
        return ""

    # "XTSE:BTO" -> "BTO"
    if ":" in s:
        s = s.split(":")[-1].strip()

    # "BTO.TO" -> "BTO"
    for suf in KNOWN_SUFFIXES:
        if s.endswith(suf):
            s = s[: -len(suf)]
            break

    # Remove whitespace/commas
    s = re.sub(r"[,\s]+", "", s)
    return s

# -----------------------------
# Loaders
# -----------------------------
@st.cache_data(show_spinner=False)
def load_table_from_upload(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return pd.read_excel(uploaded_file)
    return pd.read_csv(uploaded_file)

@st.cache_data(show_spinner=False)
def load_base(base_path: str, uploaded_file=None) -> pd.DataFrame:
    if uploaded_file is not None:
        return load_table_from_upload(uploaded_file)

    p = Path(base_path)
    if not p.exists():
        raise FileNotFoundError(f"Base file not found: {base_path}")

    if p.suffix.lower() in [".xlsx", ".xls"]:
        return pd.read_excel(p)
    return pd.read_csv(p)

def build_symbol_index(df: pd.DataFrame) -> pd.DataFrame:
    """Add helper keys for matching."""
    out = df.copy()
    symbol_col = "Symbol" if "Symbol" in out.columns else out.columns[0]
    out["__ticker_key__"] = out[symbol_col].astype(str).apply(normalize_ticker)
    out["__symbol_key__"] = out[symbol_col].astype(str).str.strip().str.upper()
    return out

def infer_meta_cols(df: pd.DataFrame) -> List[str]:
    meta = [c for c in META_COLS_DEFAULT if c in df.columns]
    # Ensure Symbol appears first if present
    if "Symbol" in df.columns and "Symbol" not in meta:
        meta = ["Symbol"] + meta
    return meta

def get_qualifier_cols(df: pd.DataFrame, meta_cols: List[str]) -> List[str]:
    ignore = set(meta_cols + ["__ticker_key__", "__symbol_key__"])
    return [c for c in df.columns if c not in ignore]

# -----------------------------
# Parsing inputs
# -----------------------------
def parse_manual_tickers(text: str) -> List[str]:
    if not text:
        return []
    parts = re.split(r"[\n,; \t]+", text)
    keys = [normalize_ticker(p) for p in parts]
    keys = [k for k in keys if k]
    seen = set()
    out = []
    for k in keys:
        if k not in seen:
            seen.add(k)
            out.append(k)
    return out

def infer_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols_lower = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        if cand in cols_lower:
            return cols_lower[cand]
    # fuzzy contains
    for c in df.columns:
        cl = str(c).strip().lower()
        for cand in candidates:
            if cand in cl:
                return c
    return None

def parse_uploaded_ticker_file(uploaded) -> List[str]:
    if uploaded is None:
        return []
    tdf = load_table_from_upload(uploaded)
    if tdf.empty:
        return []

    col = infer_column(tdf, ["ticker", "tickers", "symbol", "symbols", "symbole", "symb"])
    if col is None:
        col = tdf.columns[0]

    keys = tdf[col].dropna().astype(str).tolist()
    keys = [normalize_ticker(x) for x in keys]
    keys = [k for k in keys if k]

    seen = set()
    out = []
    for k in keys:
        if k not in seen:
            seen.add(k)
            out.append(k)
    return out

# -----------------------------
# Core matching
# -----------------------------
def find_memberships(
    df_indexed: pd.DataFrame,
    ticker_keys: List[str],
    meta_cols: List[str],
    qualifier_cols: List[str],
    match_mode: str = "Ticker",
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Returns:
      - long_results: one row per input ticker (plus per matched row if duplicates exist)
      - matrix: numeric/boolean matrix (tickers x qualifiers) aggregated across matches
    """
    long_rows = []
    if match_mode == "Exact Symbol":
        key_col = "__symbol_key__"
        wanted = [str(k).strip().upper() for k in ticker_keys]
    else:
        key_col = "__ticker_key__"
        wanted = ticker_keys

    matrix_rows: Dict[str, Dict[str, float]] = {}

    for input_key in wanted:
        if not input_key:
            continue

        matches = df_indexed[df_indexed[key_col] == input_key]
        if matches.empty:
            long_rows.append({
                "Input": input_key,
                "Found?": False,
                "Matched Symbol": "",
                "Membership Columns": "",
                "Membership Count": 0,
            })
            continue

        matrix_rows.setdefault(input_key, {})
        for _, r in matches.iterrows():
            memberships = []
            for c in qualifier_cols:
                v = r.get(c, None)
                if pd.notna(v) and str(v).strip() != "":
                    memberships.append((c, v))
                    try:
                        matrix_rows[input_key][c] = max(matrix_rows[input_key].get(c, 0), float(v))
                    except Exception:
                        matrix_rows[input_key][c] = 1

            mem_cols = ", ".join([m[0] for m in memberships])
            row = {
                "Input": input_key,
                "Found?": True,
                "Matched Symbol": r.get("Symbol", ""),
                "Membership Columns": mem_cols,
                "Membership Count": len(memberships),
            }
            for mc in meta_cols:
                if mc in matches.columns:
                    row[mc] = r.get(mc, "")
            long_rows.append(row)

    long_results = pd.DataFrame(long_rows)

    matrix = pd.DataFrame.from_dict(matrix_rows, orient="index")
    matrix.index.name = "Input"
    matrix = matrix.fillna(0)

    return long_results, matrix

# -----------------------------
# Portfolio helpers
# -----------------------------
def clean_weight_series(s: pd.Series) -> pd.Series:
    """Parse weights like '4.25' or '4.25%' or 0.0425 into percent."""
    s2 = s.copy()
    # Remove percent signs and commas
    s2 = s2.astype(str).str.replace("%", "", regex=False).str.replace(",", "", regex=False)
    s2 = pd.to_numeric(s2, errors="coerce")
    # If mostly <= 1, assume decimals -> convert to %
    if s2.dropna().empty:
        return s2
    frac_share = (s2.dropna() <= 1).mean()
    if frac_share > 0.7:
        s2 = s2 * 100.0
    return s2

def portfolio_column_exposure(found: pd.DataFrame, qualifier_cols: List[str], weight_col: str) -> pd.DataFrame:
    rows = []
    for c in qualifier_cols:
        m = found[c].notna() & (found[c].astype(str).str.strip() != "")
        rows.append({
            "Column": c,
            "Holdings (count)": int(m.sum()),
            "Weight (%)": float(found.loc[m, weight_col].sum()),
        })
    return pd.DataFrame(rows).sort_values("Weight (%)", ascending=False)

def portfolio_sector_exposure(found: pd.DataFrame, weight_col: str) -> pd.DataFrame:
    if "Sector" not in found.columns:
        return pd.DataFrame(columns=["Sector", "Holdings", "Weight (%)"])
    out = (
        found.groupby("Sector", dropna=False)
        .agg(Holdings=("Input", "count"), **{"Weight (%)": (weight_col, "sum")})
        .sort_values("Weight (%)", ascending=False)
        .reset_index()
    )
    return out

# -----------------------------
# Exports
# -----------------------------
def to_excel_bytes(
    results_df: pd.DataFrame,
    matrix_df: pd.DataFrame,
    extra_sheets: Optional[Dict[str, pd.DataFrame]] = None,
) -> bytes:
    import openpyxl
    from openpyxl.utils import get_column_letter

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        results_df.to_excel(writer, index=False, sheet_name="Results")
        matrix_df.to_excel(writer, index=True, sheet_name="Matrix")

        if extra_sheets:
            for name, df in extra_sheets.items():
                df.to_excel(writer, index=False, sheet_name=name[:31])

        # Auto-fit widths (cap at 60)
        for sheet_name in writer.book.sheetnames:
            ws = writer.book[sheet_name]
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    val = "" if cell.value is None else str(cell.value)
                    if len(val) > max_len:
                        max_len = len(val)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    return bio.getvalue()

# -----------------------------
# Charts (matplotlib)
# -----------------------------
def chart_pie(values: List[float], labels: List[str], title: str):
    fig, ax = plt.subplots(figsize=(5.2, 3.6))
    ax.pie(values, labels=labels, autopct=lambda p: f"{p:.1f}%" if p > 0 else "", startangle=90)
    ax.set_title(title)
    st.pyplot(fig, clear_figure=True)

def chart_barh(df: pd.DataFrame, x_col: str, y_col: str, title: str, top_n: int = 12):
    d = df.head(top_n).iloc[::-1]  # reverse for horizontal bar
    fig, ax = plt.subplots(figsize=(7.2, max(3.6, 0.35 * len(d) + 1.2)))
    ax.barh(d[x_col].astype(str), d[y_col].astype(float))
    ax.set_title(title)
    ax.set_xlabel(y_col)
    st.pyplot(fig, clear_figure=True)

def chart_hist(values: pd.Series, title: str, bins: int = 10, xlabel: str = ""):
    fig, ax = plt.subplots(figsize=(6.2, 3.6))
    ax.hist(values.dropna().astype(float), bins=bins)
    ax.set_title(title)
    if xlabel:
        ax.set_xlabel(xlabel)
    st.pyplot(fig, clear_figure=True)

# -----------------------------
# UI
# -----------------------------
st.title("Cross-Qualifying Ticker Finder")
st.caption("Cross-reference a ticker list (or a weighted portfolio) to see which columns each ticker belongs to — based on your base file.")

with st.sidebar:
    st.header("1) Base file")
    st.write("Upload your base Excel/CSV (optional). If you don’t upload, the app uses the bundled file path below.")
    base_upload = st.file_uploader("Upload base file (.csv, .xlsx)", type=["csv", "xlsx", "xls"], key="base_upload")
    base_path = st.text_input("Bundled base path", value="Cross Qualifying Stocks.csv")

    st.divider()

    st.header("2) Mode")
    mode = st.radio("Choose input type", ["Ticker list", "Portfolio (with weights)"], index=0)

    st.divider()

    st.header("3) Options")
    match_mode = st.radio(
        "Match using",
        ["Ticker", "Exact Symbol"],
        index=0,
        help="Ticker mode maps 'BTO' to 'XTSE:BTO'. Exact Symbol expects full symbols like 'XTSE:BTO'.",
    )
    show_only_found = st.toggle("Show only found tickers", value=False)
    show_matrix_values = st.toggle("Matrix: show numeric values (when available)", value=True)

# Load base
try:
    base_df = load_base(base_path, uploaded_file=base_upload)
except Exception as e:
    st.error(f"Could not load base file. Details: {e}")
    st.stop()

base_df = build_symbol_index(base_df)
meta_cols = infer_meta_cols(base_df)
qualifier_cols = get_qualifier_cols(base_df, meta_cols)

# Input tickers
ticker_keys: List[str] = []
portfolio_df = None
weight_col = None

if mode == "Ticker list":
    with st.sidebar:
        st.header("4) Input tickers")
        manual = st.text_area(
            "Paste tickers (comma/newline separated)",
            placeholder="Example:\nBTO\nOTEX\nEFX\nor: XTSE:BTO, XTSE:OTEX",
            height=140,
        )
        tickers_upload = st.file_uploader("…or upload a ticker list (.csv, .xlsx)", type=["csv", "xlsx", "xls"], key="tickers_upload")

    ticker_keys.extend(parse_manual_tickers(manual))
    ticker_keys.extend(parse_uploaded_ticker_file(tickers_upload))

else:
    with st.sidebar:
        st.header("4) Upload portfolio")
        port_upload = st.file_uploader("Upload portfolio (.csv, .xlsx)", type=["csv", "xlsx", "xls"], key="port_upload")
        if port_upload is not None:
            portfolio_df = load_table_from_upload(port_upload)

            # Let user pick columns (with smart defaults)
            ticker_guess = infer_column(portfolio_df, ["ticker", "symbol", "symbole", "ric"])
            weight_guess = infer_column(portfolio_df, ["weight", "cible", "allocation", "target", "%"])

            ticker_col = st.selectbox(
                "Ticker column",
                options=list(portfolio_df.columns),
                index=(list(portfolio_df.columns).index(ticker_guess) if ticker_guess in list(portfolio_df.columns) else 0),
            )
            weight_col = st.selectbox(
                "Weight column (%)",
                options=list(portfolio_df.columns),
                index=(list(portfolio_df.columns).index(weight_guess) if weight_guess in list(portfolio_df.columns) else min(1, len(portfolio_df.columns)-1)),
                help="Accepts values like 4.25 or 4.25% or 0.0425 (auto-converts).",
            )

            portfolio_df = portfolio_df.copy()
            portfolio_df["__ticker_key__"] = portfolio_df[ticker_col].astype(str).apply(normalize_ticker)
            portfolio_df["__weight_pct__"] = clean_weight_series(portfolio_df[weight_col])

            ticker_keys = portfolio_df["__ticker_key__"].dropna().astype(str).tolist()
            ticker_keys = [t for t in ticker_keys if t]

# De-dupe tickers preserving order
seen = set()
ticker_keys = [t for t in ticker_keys if not (t in seen or seen.add(t))]

# Overview
col1, col2, col3 = st.columns([1.3, 1, 1])

with col1:
    st.subheader("Base dataset")
    st.write(f"Rows: **{len(base_df):,}** · Qualifier columns: **{len(qualifier_cols)}**")
    with st.expander("Preview base file"):
        preview_cols = [c for c in (meta_cols + qualifier_cols[:10]) if c in base_df.columns]
        st.dataframe(base_df[preview_cols].head(25), use_container_width=True)

with col2:
    st.subheader("Your input")
    st.write(f"Tickers provided: **{len(ticker_keys)}**")
    if ticker_keys:
        st.code("\n".join(ticker_keys[:30]) + ("\n…" if len(ticker_keys) > 30 else ""))

with col3:
    st.subheader("How membership works")
    st.write("Any non-empty value in a qualifier column counts as membership.")
    st.write("If a cell has a number (e.g., 2 or 7), the matrix can display it (or just show 0/1).")

st.divider()
st.subheader("Results")

if not ticker_keys:
    st.info("Add tickers (paste/upload) or upload a portfolio to see results.")
    st.stop()

long_results, matrix = find_memberships(
    df_indexed=base_df,
    ticker_keys=ticker_keys,
    meta_cols=meta_cols,
    qualifier_cols=qualifier_cols,
    match_mode=match_mode,
)

if show_only_found:
    long_results = long_results[long_results["Found?"] == True].copy()

# If in portfolio mode, attach weights
extra_sheets = {}
if mode == "Portfolio (with weights)" and portfolio_df is not None:
    # Build weight map
    wmap = (
        portfolio_df.dropna(subset=["__ticker_key__"])
        .groupby("__ticker_key__", as_index=False)["__weight_pct__"]
        .sum()
        .rename(columns={"__ticker_key__": "Input", "__weight_pct__": "Weight (%)"})
    )
    long_results = long_results.merge(wmap, on="Input", how="left")

# Show long results
st.dataframe(long_results, use_container_width=True, hide_index=True)

# Summary counts
st.divider()
st.subheader("Summary (how many of your tickers appear in each column)")

presence = (matrix > 0).astype(int)
summary = presence.sum(axis=0).sort_values(ascending=False).rename("Count").to_frame()
st.dataframe(summary, use_container_width=True)

# Matrix view
st.divider()
st.subheader("Matrix view (tickers × columns)")
matrix_view = matrix.copy()
if not show_matrix_values:
    matrix_view = (matrix_view > 0).astype(int)
st.dataframe(matrix_view, use_container_width=True)

# -----------------------------
# Portfolio Insights (charts)
# -----------------------------
if mode == "Portfolio (with weights)" and portfolio_df is not None:
    st.divider()
    st.header("Portfolio insights (weighted)")

    # Determine mapped/unmapped by merging matrix index to weights
    weights = (
        portfolio_df.dropna(subset=["__ticker_key__"])
        .groupby("__ticker_key__", as_index=False)["__weight_pct__"]
        .sum()
        .rename(columns={"__ticker_key__": "Input", "__weight_pct__": "Weight (%)"})
    )

    mapped_inputs = set(matrix.index.tolist())
    weights["Mapped?"] = weights["Input"].apply(lambda x: x in mapped_inputs)

    mapped_w = float(weights.loc[weights["Mapped?"], "Weight (%)"].sum())
    unmapped_w = float(weights.loc[~weights["Mapped?"], "Weight (%)"].sum())
    total_w = float(weights["Weight (%)"].sum()) if not weights.empty else 0.0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total weight (%)", f"{total_w:.2f}")
    m2.metric("Mapped weight (%)", f"{mapped_w:.2f}")
    m3.metric("Unmapped weight (%)", f"{unmapped_w:.2f}")
    m4.metric("Coverage (%)", f"{(100*mapped_w/total_w if total_w else 0):.1f}")

    c1, c2 = st.columns([1, 1.2])
    with c1:
        chart_pie([mapped_w, unmapped_w], ["Mapped", "Unmapped"], "Coverage by weight")
    with c2:
        chart_hist(long_results.loc[long_results["Found?"] == True, "Membership Count"], "Membership Count distribution (mapped holdings)", bins=10, xlabel="Membership Count")

    # Build a 'found' table with meta + weights (one row per input, so merge on Input and choose first match)
    found_rows = long_results[long_results["Found?"] == True].copy()
    if "Weight (%)" not in found_rows.columns:
        found_rows = found_rows.merge(weights[["Input", "Weight (%)"]], on="Input", how="left")

    # Column exposure by weight (mapped portion)
    # Need base values for each input; use matrix_view (numeric) and weights
    if not matrix.empty and not weights.empty:
        mw = matrix.copy()
        mw = mw.merge(weights.set_index("Input")[["Weight (%)"]], left_index=True, right_index=True, how="left")
        # For each qualifier, sum weights where membership > 0
        exp_rows = []
        for c in matrix.columns:
            m = mw[c] > 0
            exp_rows.append({"Column": c, "Weight (%)": float(mw.loc[m, "Weight (%)"].sum()), "Holdings (count)": int(m.sum())})
        col_exp = pd.DataFrame(exp_rows).sort_values("Weight (%)", ascending=False)

        st.subheader("Top column exposures (by weight, mapped portion)")
        e1, e2 = st.columns([1.1, 1])
        with e1:
            st.dataframe(col_exp, use_container_width=True)
        with e2:
            chart_barh(col_exp, "Column", "Weight (%)", "Top columns by weight", top_n=12)

        extra_sheets["Column Exposure"] = col_exp

    # Sector exposure (only for mapped)
    if "Sector" in found_rows.columns and "Weight (%)" in found_rows.columns:
        sector_exp = (
            found_rows.groupby("Sector", dropna=False)
            .agg(**{"Holdings (count)": ("Input", "nunique"), "Weight (%)": ("Weight (%)", "sum")})
            .sort_values("Weight (%)", ascending=False)
            .reset_index()
        )
        st.subheader("Sector exposure (mapped portion)")
        s1, s2 = st.columns([1.1, 1])
        with s1:
            st.dataframe(sector_exp, use_container_width=True)
        with s2:
            chart_barh(sector_exp.rename(columns={"Sector":"Column"}), "Column", "Weight (%)", "Top sectors by weight", top_n=12)

        extra_sheets["Sector Exposure"] = sector_exp

    # Top cross-qualifiers by (Membership Count, weight)
    if "Membership Count" in found_rows.columns:
        top_cross = found_rows.sort_values(["Membership Count", "Weight (%)"], ascending=[False, False]).head(15)
        keep_cols = [c for c in ["Input", "Matched Symbol", "Sector", "Weight (%)", "Membership Count", "Membership Columns"] if c in top_cross.columns]
        st.subheader("Top cross-qualifiers (mapped holdings)")
        st.dataframe(top_cross[keep_cols], use_container_width=True)
        extra_sheets["Top Cross-Qualifiers"] = top_cross[keep_cols]

    # Unmapped list
    unmapped = weights.loc[~weights["Mapped?"]].sort_values("Weight (%)", ascending=False)
    st.subheader("Unmapped tickers (not found in base file)")
    st.dataframe(unmapped, use_container_width=True, hide_index=True)
    extra_sheets["Unmapped (Portfolio)"] = unmapped

# -----------------------------
# Downloads
# -----------------------------
st.divider()
st.subheader("Download")

csv_bytes = long_results.to_csv(index=False).encode("utf-8")
st.download_button("Download results (CSV)", data=csv_bytes, file_name="cross_qualifying_results.csv", mime="text/csv")

xlsx_bytes = to_excel_bytes(long_results, matrix_view, extra_sheets=extra_sheets if extra_sheets else None)
st.download_button("Download results (Excel)", data=xlsx_bytes, file_name="cross_qualifying_results.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
'''

reqs = "streamlit\npandas\nopenpyxl\nmatplotlib\n"

app_path = "/mnt/data/app.py"
req_path = "/mnt/data/requirements.txt"
with open(app_path, "w", encoding="utf-8") as f:
    f.write(app_code.strip() + "\n")
with open(req_path, "w", encoding="utf-8") as f:
    f.write(reqs)

app_path, req_path, os.path.getsize(app_path), os.path.getsize(req_path)

